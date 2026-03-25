import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

HDR_FILL   = PatternFill("solid", start_color="1F3864")
HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
SEV_COLORS = {"Critical": "FF0000", "High": "FF6600", "Medium": "FFC000"}
THIN       = Side(style="thin", color="CCCCCC")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _style_header(ws, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = 28

def _write_df(ws, df: pd.DataFrame):
    headers = list(df.columns)
    _style_header(ws, headers)
    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, (h, val) in enumerate(zip(headers, row), 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=(c > 7))
            if h == "Vuln Severity":
                cell.font = Font(name="Arial", size=9, bold=True,
                                 color=SEV_COLORS.get(str(val), "000000"))
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 14 if i < 8 else 40
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def write_output(path: str, groups: dict, sol_map: dict):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    # One sheet per (server, class) group
    for (server, cls), df in sorted(groups.items()):
        sheet_name = f"{server}_{cls}"[:31]
        ws = wb.create_sheet(title=sheet_name)
        _write_df(ws, df)

    # Final sheet: Vuln Solution → Server IPs mapping
    ws_sol = wb.create_sheet(title="Sol_to_IPs")
    ws_sol.cell(1, 1, "Vulnerability Solution").font = HDR_FONT
    ws_sol.cell(1, 1).fill = HDR_FILL
    ws_sol.cell(1, 2, "Affected Server IPs").font = HDR_FONT
    ws_sol.cell(1, 2).fill = HDR_FILL

    for r, (sol, ips) in enumerate(sorted(sol_map.items()), 2):
        ws_sol.cell(r, 1, sol).border = BORDER
        ws_sol.cell(r, 2, ", ".join(ips)).border = BORDER

    ws_sol.column_dimensions["A"].width = 60
    ws_sol.column_dimensions["B"].width = 50

    wb.save(path)